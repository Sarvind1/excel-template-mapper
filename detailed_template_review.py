import openpyxl
from openpyxl.utils import get_column_letter
import re

# Load the Excel file
file_path = 'YW1 Inbound PL.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=False)

# Get source data headers
yw1_sheet = wb['YW1 Inbound PL']
container_sheet = wb['Container']

# Build header mappings
yw1_headers = {}
for col in range(1, yw1_sheet.max_column + 1):
    cell = yw1_sheet.cell(1, col)
    if cell.value:
        col_letter = get_column_letter(col)
        yw1_headers[str(cell.value).strip()] = col_letter

container_headers = {}
for col in range(1, container_sheet.max_column + 1):
    cell = container_sheet.cell(1, col)
    if cell.value:
        col_letter = get_column_letter(col)
        container_headers[str(cell.value).strip()] = col_letter

print("="*80)
print("DETAILED SCI TEMPLATE STRUCTURE")
print("="*80)

sci_sheet = wb['SCI Template - Single Batch']

# Find the table structure - look for headers/field names
print("\nScanning SCI Template structure...")
print(f"Dimensions: {sci_sheet.max_row} rows x {sci_sheet.max_column} columns")

# Sample first 50 rows to understand structure
print("\nFirst 50 rows sample:")
for row in range(1, min(51, sci_sheet.max_row + 1)):
    row_data = []
    for col in range(1, min(15, sci_sheet.max_column + 1)):  # First 15 columns
        cell = sci_sheet.cell(row, col)
        if cell.value:
            cell_ref = f"{get_column_letter(col)}{row}"
            if str(cell.value).startswith('='):
                row_data.append(f"{cell_ref}:[FORMULA]")
            else:
                val_preview = str(cell.value)[:50]
                row_data.append(f"{cell_ref}:{val_preview}")

    if row_data:
        print(f"Row {row}: {' | '.join(row_data)}")

print("\n" + "="*80)
print("DETAILED PL TEMPLATE STRUCTURE")
print("="*80)

pl_sheet = wb['PL Template - Single Batch']
print(f"\nDimensions: {pl_sheet.max_row} rows x {pl_sheet.max_column} columns")

print("\nFirst 50 rows sample:")
for row in range(1, min(51, pl_sheet.max_row + 1)):
    row_data = []
    for col in range(1, min(15, pl_sheet.max_column + 1)):
        cell = pl_sheet.cell(row, col)
        if cell.value:
            cell_ref = f"{get_column_letter(col)}{row}"
            if str(cell.value).startswith('='):
                row_data.append(f"{cell_ref}:[FORMULA]")
            else:
                val_preview = str(cell.value)[:50]
                row_data.append(f"{cell_ref}:{val_preview}")

    if row_data:
        print(f"Row {row}: {' | '.join(row_data)}")

import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter

# Load the Excel file
file_path = 'YW1 Inbound PL.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=False)

print("Available sheets in the workbook:")
print("="*50)
for sheet_name in wb.sheetnames:
    print(f"- {sheet_name}")

print("\n" + "="*80)
print("ANALYZING: SCI Template - Single Batch")
print("="*80)

sci_sheet = wb['SCI Template - Single Batch']
print(f"\nDimensions: {sci_sheet.dimensions}")
print(f"Max Row: {sci_sheet.max_row}, Max Column: {sci_sheet.max_column}")

# Get all non-empty cells and their formulas
print("\n--- Fields and Formulas in SCI Template ---")
for row in sci_sheet.iter_rows(min_row=1, max_row=sci_sheet.max_row):
    for cell in row:
        if cell.value is not None:
            cell_ref = f"{get_column_letter(cell.column)}{cell.row}"
            if hasattr(cell, 'value') and str(cell.value).startswith('='):
                print(f"{cell_ref}: FORMULA = {cell.value}")
            else:
                # Show labels/headers
                if isinstance(cell.value, str) and len(cell.value) < 100:
                    print(f"{cell_ref}: {cell.value}")

print("\n" + "="*80)
print("ANALYZING: PL Template - Single Batch")
print("="*80)

pl_sheet = wb['PL Template - Single Batch']
print(f"\nDimensions: {pl_sheet.dimensions}")
print(f"Max Row: {pl_sheet.max_row}, Max Column: {pl_sheet.max_column}")

print("\n--- Fields and Formulas in PL Template ---")
for row in pl_sheet.iter_rows(min_row=1, max_row=pl_sheet.max_row):
    for cell in row:
        if cell.value is not None:
            cell_ref = f"{get_column_letter(cell.column)}{cell.row}"
            if hasattr(cell, 'value') and str(cell.value).startswith('='):
                print(f"{cell_ref}: FORMULA = {cell.value}")
            else:
                # Show labels/headers
                if isinstance(cell.value, str) and len(cell.value) < 100:
                    print(f"{cell_ref}: {cell.value}")

print("\n" + "="*80)
print("ANALYZING: YW1 Inbound PL (Source Data)")
print("="*80)

yw1_sheet = wb['YW1 Inbound PL']
print(f"\nDimensions: {yw1_sheet.dimensions}")
print(f"Max Row: {yw1_sheet.max_row}, Max Column: {yw1_sheet.max_column}")

# Get headers
print("\n--- Column Headers in YW1 Inbound PL ---")
headers = []
for col in range(1, yw1_sheet.max_column + 1):
    cell = yw1_sheet.cell(1, col)
    if cell.value:
        headers.append(f"{get_column_letter(col)}: {cell.value}")
        print(f"{get_column_letter(col)}: {cell.value}")

print("\n" + "="*80)
print("ANALYZING: Container (Source Data)")
print("="*80)

container_sheet = wb['Container']
print(f"\nDimensions: {container_sheet.dimensions}")
print(f"Max Row: {container_sheet.max_row}, Max Column: {container_sheet.max_column}")

print("\n--- Column Headers in Container ---")
for col in range(1, container_sheet.max_column + 1):
    cell = container_sheet.cell(1, col)
    if cell.value:
        print(f"{get_column_letter(col)}: {cell.value}")

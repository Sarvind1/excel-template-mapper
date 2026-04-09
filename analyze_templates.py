import openpyxl
from openpyxl.utils import get_column_letter
import json

# Load the Excel file
file_path = 'YW1 Inbound PL.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=False)

# Get source data headers
yw1_sheet = wb['YW1 Inbound PL']
container_sheet = wb['Container']

print("="*80)
print("SOURCE DATA HEADERS")
print("="*80)

print("\nYW1 Inbound PL Headers:")
yw1_headers = {}
for col in range(1, yw1_sheet.max_column + 1):
    cell = yw1_sheet.cell(1, col)
    if cell.value:
        col_letter = get_column_letter(col)
        yw1_headers[col_letter] = str(cell.value)
        print(f"  {col_letter}: {cell.value}")

print("\nContainer Headers:")
container_headers = {}
for col in range(1, container_sheet.max_column + 1):
    cell = container_sheet.cell(1, col)
    if cell.value:
        col_letter = get_column_letter(col)
        container_headers[col_letter] = str(cell.value)
        print(f"  {col_letter}: {cell.value}")

# Analyze SCI Template
print("\n" + "="*80)
print("SCI TEMPLATE - Single Batch ANALYSIS")
print("="*80)

sci_sheet = wb['SCI Template - Single Batch']
sci_fields = {}

# Scan for field labels and values
for row in range(1, min(sci_sheet.max_row + 1, 200)):  # Limit to first 200 rows
    for col in range(1, min(sci_sheet.max_column + 1, 50)):  # Limit to first 50 cols
        cell = sci_sheet.cell(row, col)

        # Check if this looks like a label (text in one cell, potentially value/formula in adjacent cell)
        if cell.value and isinstance(cell.value, str) and ':' in str(cell.value):
            label = cell.value
            value_cell = sci_sheet.cell(row, col + 1)
            cell_ref = f"{get_column_letter(col + 1)}{row}"

            if value_cell.value:
                if str(value_cell.value).startswith('='):
                    sci_fields[label] = {
                        'cell': cell_ref,
                        'type': 'formula',
                        'value': value_cell.value
                    }
                else:
                    sci_fields[label] = {
                        'cell': cell_ref,
                        'type': 'static',
                        'value': value_cell.value
                    }
            else:
                sci_fields[label] = {
                    'cell': cell_ref,
                    'type': 'empty',
                    'value': None
                }

print(f"\nFound {len(sci_fields)} labeled fields in SCI Template")
print("\nSample fields:")
for i, (label, info) in enumerate(list(sci_fields.items())[:10]):
    print(f"  {label} -> {info['cell']} ({info['type']})")

# Analyze PL Template
print("\n" + "="*80)
print("PL TEMPLATE - Single Batch ANALYSIS")
print("="*80)

pl_sheet = wb['PL Template - Single Batch']
pl_fields = {}

# Scan for field labels and values
for row in range(1, min(pl_sheet.max_row + 1, 200)):
    for col in range(1, min(pl_sheet.max_column + 1, 50)):
        cell = pl_sheet.cell(row, col)

        if cell.value and isinstance(cell.value, str) and ':' in str(cell.value):
            label = cell.value
            value_cell = pl_sheet.cell(row, col + 1)
            cell_ref = f"{get_column_letter(col + 1)}{row}"

            if value_cell.value:
                if str(value_cell.value).startswith('='):
                    pl_fields[label] = {
                        'cell': cell_ref,
                        'type': 'formula',
                        'value': value_cell.value
                    }
                else:
                    pl_fields[label] = {
                        'cell': cell_ref,
                        'type': 'static',
                        'value': value_cell.value
                    }
            else:
                pl_fields[label] = {
                    'cell': cell_ref,
                    'type': 'empty',
                    'value': None
                }

print(f"\nFound {len(pl_fields)} labeled fields in PL Template")
print("\nSample fields:")
for i, (label, info) in enumerate(list(pl_fields.items())[:10]):
    print(f"  {label} -> {info['cell']} ({info['type']})")

# Save detailed analysis to JSON
analysis = {
    'source_data': {
        'yw1_headers': yw1_headers,
        'container_headers': container_headers
    },
    'sci_template': sci_fields,
    'pl_template': pl_fields
}

with open('template_analysis.json', 'w') as f:
    json.dump(analysis, f, indent=2)

print("\n" + "="*80)
print("Detailed analysis saved to template_analysis.json")
print("="*80)

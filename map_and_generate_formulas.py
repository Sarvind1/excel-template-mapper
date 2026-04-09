import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
import re

# Load the Excel file
file_path = 'YW1 Inbound PL.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=False)

# Source sheets
yw1_sheet = wb['YW1 Inbound PL']
container_sheet = wb['Container']

# Template sheets
sci_sheet = wb['SCI Template - Single Batch']
pl_sheet = wb['PL Template - Single Batch']

# Build comprehensive mappings
print("="*80)
print("MAPPING SOURCE DATA TO TEMPLATE FIELDS")
print("="*80)

# YW1 Inbound PL headers
yw1_mapping = {}
for col in range(1, yw1_sheet.max_column + 1):
    header = yw1_sheet.cell(1, col).value
    if header:
        yw1_mapping[str(header).strip()] = get_column_letter(col)

# Container headers
container_mapping = {}
for col in range(1, container_sheet.max_column + 1):
    header = container_sheet.cell(1, col).value
    if header:
        container_mapping[str(header).strip()] = get_column_letter(col)

print("\n--- Available YW1 Inbound PL columns ---")
for key in sorted(yw1_mapping.keys()):
    print(f"  {key}: Column {yw1_mapping[key]}")

print("\n--- Available Container columns ---")
for key in sorted(container_mapping.keys()):
    print(f"  {key}: Column {container_mapping[key]}")

# Analyze SCI Template fields
print("\n" + "="*80)
print("SCI TEMPLATE FIELD ANALYSIS")
print("="*80)

sci_fields_to_map = []
sci_unmappable_fields = []

# Extract header row and data rows
sci_header_row = 22
sci_data_start_row = 23

# Get headers from row 22
sci_headers = []
for col in range(1, sci_sheet.max_column + 1):
    header_cell = sci_sheet.cell(sci_header_row, col)
    if header_cell.value:
        sci_headers.append({
            'col': col,
            'col_letter': get_column_letter(col),
            'name': str(header_cell.value).strip(),
            'cell_ref': f"{get_column_letter(col)}{sci_header_row}"
        })

print(f"\nFound {len(sci_headers)} columns in SCI Template:")
for h in sci_headers:
    print(f"  {h['col_letter']}: {h['name']}")

# Map each SCI field
sci_field_mapping = []
for h in sci_headers:
    field_name = h['name']
    mapped = False
    source_sheet = None
    source_column = None
    formula = None

    # Try to map to YW1 or Container
    if field_name == 'PO#':
        if 'PO' in yw1_mapping:
            source_sheet = 'YW1 Inbound PL'
            source_column = yw1_mapping['PO']
            mapped = True
    elif field_name == 'Batch ID':
        if 'Batch' in yw1_mapping:
            source_sheet = 'YW1 Inbound PL'
            source_column = yw1_mapping['Batch']
            mapped = True
    elif field_name == 'Razin' or field_name == 'Product name':
        # Razin might map to ASIN or IC SKU
        if 'ASIN' in yw1_mapping:
            source_sheet = 'YW1 Inbound PL'
            source_column = yw1_mapping['ASIN']
            mapped = True
    elif 'Product name' in field_name or field_name == 'Product name':
        if 'Description' in yw1_mapping:
            source_sheet = 'YW1 Inbound PL'
            source_column = yw1_mapping['Description']
            mapped = True
    elif field_name == 'Composition':
        if 'Material' in yw1_mapping:
            source_sheet = 'YW1 Inbound PL'
            source_column = yw1_mapping['Material']
            mapped = True
    elif field_name == 'HS code':
        # HS code might not be directly available
        mapped = False
    elif 'Quantity' in field_name and 'Units' in field_name:
        if 'Qty' in yw1_mapping:
            source_sheet = 'YW1 Inbound PL'
            source_column = yw1_mapping['Qty']
            mapped = True
    elif 'Carton quantity' in field_name:
        if 'Cartons' in yw1_mapping:
            source_sheet = 'YW1 Inbound PL'
            source_column = yw1_mapping['Cartons']
            mapped = True
    elif 'Parcels per MC' in field_name:
        if 'Qty/Ctn' in yw1_mapping:
            source_sheet = 'YW1 Inbound PL'
            source_column = yw1_mapping['Qty/Ctn']
            mapped = True
    elif 'Carton Vol' in field_name or 'CBM' in field_name:
        if 'CBM' in yw1_mapping:
            source_sheet = 'YW1 Inbound PL'
            source_column = yw1_mapping['CBM']
            mapped = True

    sci_field_mapping.append({
        'field_name': field_name,
        'template_column': h['col_letter'],
        'mapped': mapped,
        'source_sheet': source_sheet,
        'source_column': source_column
    })

# Analyze PL Template fields
print("\n" + "="*80)
print("PL TEMPLATE FIELD ANALYSIS")
print("="*80)

pl_header_row = 22
pl_data_start_row = 23

pl_headers = []
for col in range(1, pl_sheet.max_column + 1):
    header_cell = pl_sheet.cell(pl_header_row, col)
    if header_cell.value:
        pl_headers.append({
            'col': col,
            'col_letter': get_column_letter(col),
            'name': str(header_cell.value).strip(),
            'cell_ref': f"{get_column_letter(col)}{pl_header_row}"
        })

print(f"\nFound {len(pl_headers)} columns in PL Template:")
for h in pl_headers:
    print(f"  {h['col_letter']}: {h['name']}")

# Map each PL field
pl_field_mapping = []
for h in pl_headers:
    field_name = h['name']
    mapped = False
    source_sheet = None
    source_column = None

    if field_name == 'Batch ID':
        if 'Batch' in yw1_mapping:
            source_sheet = 'YW1 Inbound PL'
            source_column = yw1_mapping['Batch']
            mapped = True
    elif field_name == 'Razin':
        if 'ASIN' in yw1_mapping:
            source_sheet = 'YW1 Inbound PL'
            source_column = yw1_mapping['ASIN']
            mapped = True
    elif field_name == 'Product name':
        if 'Description' in yw1_mapping:
            source_sheet = 'YW1 Inbound PL'
            source_column = yw1_mapping['Description']
            mapped = True
    elif 'Product Package Dim' in field_name:
        # Need to calculate from Length x Height x Width
        if all(k in yw1_mapping for k in ['Length', 'Width', 'Height']):
            source_sheet = 'YW1 Inbound PL'
            source_column = 'CALCULATED'
            mapped = True
    elif 'Product Package Weight' in field_name:
        # Might map to Ctn Weight divided by Qty/Ctn
        if 'Ctn Weight' in yw1_mapping:
            source_sheet = 'YW1 Inbound PL'
            source_column = yw1_mapping['Ctn Weight']
            mapped = True
    elif field_name == 'HS code':
        mapped = False
    elif 'Number of Cartons' in field_name:
        if 'Cartons' in yw1_mapping:
            source_sheet = 'YW1 Inbound PL'
            source_column = yw1_mapping['Cartons']
            mapped = True
    elif 'Parcels per MC' in field_name:
        if 'Qty/Ctn' in yw1_mapping:
            source_sheet = 'YW1 Inbound PL'
            source_column = yw1_mapping['Qty/Ctn']
            mapped = True
    elif 'Quantity' in field_name and 'Units' in field_name:
        if 'Qty' in yw1_mapping:
            source_sheet = 'YW1 Inbound PL'
            source_column = yw1_mapping['Qty']
            mapped = True
    elif 'Net Weight' in field_name:
        # Could be calculated
        mapped = False
    elif 'Gross Weight' in field_name:
        if 'Ctn Weight' in yw1_mapping:
            source_sheet = 'YW1 Inbound PL'
            source_column = yw1_mapping['Ctn Weight']
            mapped = True
    elif 'Master Car. Dim' in field_name:
        if all(k in yw1_mapping for k in ['Length', 'Width', 'Height']):
            source_sheet = 'YW1 Inbound PL'
            source_column = 'CALCULATED'
            mapped = True

    pl_field_mapping.append({
        'field_name': field_name,
        'template_column': h['col_letter'],
        'mapped': mapped,
        'source_sheet': source_sheet,
        'source_column': source_column
    })

# Print mapping results
print("\n" + "="*80)
print("SCI TEMPLATE MAPPING RESULTS")
print("="*80)
for mapping in sci_field_mapping:
    status = "✓ MAPPED" if mapping['mapped'] else "✗ UNMAPPED"
    if mapping['mapped']:
        print(f"{status}: {mapping['field_name']} -> {mapping['source_sheet']}!{mapping['source_column']}")
    else:
        print(f"{status}: {mapping['field_name']}")

print("\n" + "="*80)
print("PL TEMPLATE MAPPING RESULTS")
print("="*80)
for mapping in pl_field_mapping:
    status = "✓ MAPPED" if mapping['mapped'] else "✗ UNMAPPED"
    if mapping['mapped']:
        print(f"{status}: {mapping['field_name']} -> {mapping['source_sheet']}!{mapping['source_column']}")
    else:
        print(f"{status}: {mapping['field_name']}")

# Count unmapped fields
sci_unmapped = [m for m in sci_field_mapping if not m['mapped']]
pl_unmapped = [m for m in pl_field_mapping if not m['mapped']]

print("\n" + "="*80)
print("SUMMARY")
print("="*80)
print(f"SCI Template: {len(sci_field_mapping) - len(sci_unmapped)}/{len(sci_field_mapping)} fields mapped")
print(f"PL Template: {len(pl_field_mapping) - len(pl_unmapped)}/{len(pl_field_mapping)} fields mapped")
print(f"\nTotal unmapped fields: {len(sci_unmapped) + len(pl_unmapped)}")

# Create a new sheet for unmapped fields
if 'Unmappable Fields' in wb.sheetnames:
    del wb['Unmappable Fields']

unmapped_sheet = wb.create_sheet('Unmappable Fields')

# Add headers
unmapped_sheet['A1'] = 'Template'
unmapped_sheet['B1'] = 'Field Name'
unmapped_sheet['C1'] = 'Template Column'
unmapped_sheet['D1'] = 'Notes'

# Style headers
for cell in ['A1', 'B1', 'C1', 'D1']:
    unmapped_sheet[cell].font = Font(bold=True)
    unmapped_sheet[cell].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

# Add SCI unmapped fields
row = 2
for m in sci_unmapped:
    unmapped_sheet[f'A{row}'] = 'SCI Template'
    unmapped_sheet[f'B{row}'] = m['field_name']
    unmapped_sheet[f'C{row}'] = m['template_column']
    unmapped_sheet[f'D{row}'] = 'No direct mapping found in source data'
    row += 1

# Add PL unmapped fields
for m in pl_unmapped:
    unmapped_sheet[f'A{row}'] = 'PL Template'
    unmapped_sheet[f'B{row}'] = m['field_name']
    unmapped_sheet[f'C{row}'] = m['template_column']
    unmapped_sheet[f'D{row}'] = 'No direct mapping found in source data'
    row += 1

# Adjust column widths
unmapped_sheet.column_dimensions['A'].width = 20
unmapped_sheet.column_dimensions['B'].width = 40
unmapped_sheet.column_dimensions['C'].width = 15
unmapped_sheet.column_dimensions['D'].width = 50

# Save the workbook
output_file = 'YW1 Inbound PL_with_unmapped.xlsx'
wb.save(output_file)

print(f"\n✓ Created 'Unmappable Fields' tab in {output_file}")
print(f"  - {len(sci_unmapped)} SCI Template unmapped fields")
print(f"  - {len(pl_unmapped)} PL Template unmapped fields")
